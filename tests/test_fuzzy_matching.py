import os
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch, MagicMock
from dotenv import load_dotenv
import difflib
import openpyxl

# Import the functions we want to test
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from coi_auditor.pdf_parser import (
    diagnose_pdf_discovery, 
    find_best_fuzzy_matches, 
    _normalize_name, 
    _normalize_name_enhanced, 
    _get_normalized_variations,
    find_coi_pdfs
)
from coi_auditor.main import run_diagnostic_subcontractor
from coi_auditor.config import load_config

def normalize(name):
    """Legacy normalize function for backward compatibility tests."""
    return ''.join(filter(str.isalnum, name)).lower() if name else ''

class TestFuzzyMatching(unittest.TestCase):
    """Test cases for fuzzy matching functionality."""
    
    @classmethod
    def setUpClass(cls):
        """Set up test environment."""
        # Load environment variables
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        dotenv_path = os.path.join(project_root, '.env')
        load_dotenv(dotenv_path=dotenv_path)
        
        cls.excel_path = os.getenv('EXCEL_FILE_PATH')
        cls.pdf_dir = os.getenv('PDF_DIRECTORY_PATH')
        
    def test_legacy_fuzzy_matching(self):
        """Test legacy fuzzy matching between Excel names and PDF filenames."""
        if not self.excel_path or not os.path.exists(self.excel_path):
            self.skipTest(f"Excel file not found at {self.excel_path}")
        if not self.pdf_dir or not os.path.isdir(self.pdf_dir):
            self.skipTest(f"PDF directory not found at {self.pdf_dir}")

        wb = openpyxl.load_workbook(self.excel_path)
        self.assertIn('SUMMARY', wb.sheetnames, "SUMMARY sheet not found in Excel file.")
        
        sheet = wb['SUMMARY']
        header_row = int(os.getenv('EXCEL_HEADER_ROW', '6'))
        name_col_header = os.getenv('EXCEL_SUBCONTRACTOR_NAME_COL', 'Name')

        headers = [cell.value for cell in sheet[header_row]]
        self.assertIn(name_col_header, headers, f"'{name_col_header}' column not found in SUMMARY sheet headers.")
        name_idx = headers.index(name_col_header) + 1
        
        last_row = sheet.max_row
        excel_names = [sheet.cell(row=r, column=name_idx).value for r in range(header_row + 1, last_row + 1)]
        excel_names = [n for n in excel_names if n]

        pdf_files = [f for f in os.listdir(self.pdf_dir) if f.lower().endswith('.pdf')]
        pdf_bases = [os.path.splitext(f)[0] for f in pdf_files]
        normalized_pdf_bases = [normalize(b) for b in pdf_bases]

        threshold = 0.8
        results = []
        for excel_name in excel_names[:5]:  # Test first 5 for speed
            norm_excel = normalize(excel_name)
            scored = [(pdf_files[i], difflib.SequenceMatcher(None, norm_excel, normalized_pdf_bases[i]).ratio())
                      for i in range(len(pdf_files))]
            scored.sort(key=lambda x: x[1], reverse=True)
            best = [s for s in scored if s[1] >= threshold]
            results.append((excel_name, best[:3]))
            print(f"Excel: '{excel_name}' | Top matches: {best[:3]}")
        
        self.assertTrue(results, "No fuzzy matching results produced.")
        print("test_legacy_fuzzy_matching PASSED")

    def test_name_normalization_functions(self):
        """Test the name normalization functions."""
        test_cases = [
            ("ABC Construction LLC", "abcconstructionllc"),
            ("S&G Siding and Gutters", "sgsidingandgutters"),
            ("31-W Insulation", "31winsulation"),
            ("Fernando Hernandez", "fernandohernandez"),
            ("X-Stream Cleaning", "xstreamcleaning"),
            ("", ""),
            ("Test & Co., Inc.", "testcoinc"),
        ]
        
        for original, expected_simple in test_cases:
            with self.subTest(original=original):
                simple_result = _normalize_name(original)
                enhanced_result = _normalize_name_enhanced(original)
                
                self.assertEqual(simple_result, expected_simple,
                               f"Simple normalization failed for '{original}': got {simple_result!r}, expected {expected_simple!r}")
                
                # Enhanced normalization should produce a result for non-empty input
                if original:
                    self.assertTrue(enhanced_result,
                                  f"Enhanced normalization returned empty for '{original}'")
                    self.assertIsInstance(enhanced_result, str,
                                        f"Enhanced normalization should return string for '{original}'")
                else:
                    # Empty input should return empty result
                    self.assertEqual(enhanced_result, "",
                                   f"Enhanced normalization should return empty string for empty input")

    def test_name_variations_generation(self):
        """Test the generation of name variations."""
        test_cases = [
            "ABC Construction LLC",
            "S&G Siding and Gutters", 
            "31-W Insulation",
            "Fernando Hernandez"
        ]
        
        for name in test_cases:
            with self.subTest(name=name):
                variations = _get_normalized_variations(name)
                self.assertIsInstance(variations, list, f"Variations should be a list for '{name}'")
                self.assertTrue(len(variations) >= 1, f"Should generate at least one variation for '{name}'")
                # All variations should be strings
                for variation in variations:
                    self.assertIsInstance(variation, str, f"All variations should be strings for '{name}'")

    def test_fuzzy_matching_with_rapidfuzz(self):
        """Test fuzzy matching using rapidfuzz if available."""
        try:
            from rapidfuzz import fuzz
            rapidfuzz_available = True
        except ImportError:
            rapidfuzz_available = False
            
        if not rapidfuzz_available:
            self.skipTest("rapidfuzz not available")
            
        target_name = "ABC Construction"
        candidate_files = [
            "ABC_Construction_2024",
            "ABC_Const_LLC_2023", 
            "XYZ_Roofing_2024",
            "ABC_Building_2024",
            "DEF_Construction_2024"
        ]
        
        matches = find_best_fuzzy_matches(
            target_name=target_name,
            candidate_files=candidate_files,
            threshold=50.0,
            max_results=5
        )
        
        self.assertIsInstance(matches, list, "Should return a list of matches")
        # Should find at least the close matches
        self.assertTrue(len(matches) >= 2, "Should find at least 2 matches above 50% threshold")
        
        # Check that results are sorted by score descending
        scores = [score for _, score in matches]
        self.assertEqual(scores, sorted(scores, reverse=True), "Results should be sorted by score descending")

    def test_diagnostic_function_basic(self):
        """Test the basic diagnostic function."""
        # Test with a mock configuration
        with patch('coi_auditor.pdf_parser.CONFIG') as mock_config:
            mock_config.get.return_value = {
                'enabled': True,
                'threshold': 75.0,
                'algorithms': ['ratio', 'partial_ratio']
            }
            
            # Test with non-existent directory
            result = diagnose_pdf_discovery(
                subcontractor_name="Test Contractor",
                pdf_directory_path="/nonexistent/path"
            )
            
            self.assertIsInstance(result, dict, "Should return a dictionary")
            self.assertIn('subcontractor_name', result, "Should include subcontractor name")
            self.assertIn('summary', result, "Should include summary")
            self.assertIn('recommendations', result, "Should include recommendations")
            self.assertEqual(result['subcontractor_name'], "Test Contractor")

    def test_diagnostic_function_with_real_directory(self):
        """Test diagnostic function with real PDF directory if available."""
        if not self.pdf_dir or not os.path.isdir(self.pdf_dir):
            self.skipTest(f"PDF directory not available: {self.pdf_dir}")
            
        # Get a real subcontractor name from Excel if available
        test_name = "Test Contractor"
        if self.excel_path and os.path.exists(self.excel_path):
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                if 'SUMMARY' in wb.sheetnames:
                    sheet = wb['SUMMARY']
                    header_row = int(os.getenv('EXCEL_HEADER_ROW', '6'))
                    name_col_header = os.getenv('EXCEL_SUBCONTRACTOR_NAME_COL', 'Name')
                    headers = [cell.value for cell in sheet[header_row]]
                    if name_col_header in headers:
                        name_idx = headers.index(name_col_header) + 1
                        # Get first non-empty name
                        for r in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
                            cell_value = sheet.cell(row=r, column=name_idx).value
                            if cell_value:
                                test_name = str(cell_value)
                                break
            except Exception:
                pass  # Use default test name
        
        result = diagnose_pdf_discovery(
            subcontractor_name=test_name,
            pdf_directory_path=self.pdf_dir
        )
        
        self.assertIsInstance(result, dict, "Should return a dictionary")
        self.assertEqual(result['subcontractor_name'], test_name)
        self.assertIn('directory_analysis', result, "Should include directory analysis")
        self.assertIn('name_analysis', result, "Should include name analysis")
        self.assertIn('summary', result, "Should include summary")
        
        # Check directory analysis
        dir_analysis = result['directory_analysis']
        self.assertTrue(dir_analysis['path_exists'], "PDF directory should exist")
        self.assertTrue(dir_analysis['is_directory'], "PDF path should be a directory")
        self.assertTrue(dir_analysis['accessible'], "PDF directory should be accessible")

    def test_diagnostic_function_with_json_output(self):
        """Test diagnostic function with JSON output file."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp_file:
            output_file = tmp_file.name
        
        try:
            result = diagnose_pdf_discovery(
                subcontractor_name="Test Contractor",
                pdf_directory_path="/nonexistent/path",
                output_file=output_file
            )
            
            # Check that the function attempted to save the file
            # (it may fail due to nonexistent directory, but should try)
            self.assertIsInstance(result, dict, "Should return a dictionary")
            
            # If the file was created successfully, verify its contents
            if os.path.exists(output_file):
                with open(output_file, 'r', encoding='utf-8') as f:
                    saved_data = json.load(f)
                self.assertEqual(saved_data['subcontractor_name'], "Test Contractor")
                
        finally:
            # Clean up
            if os.path.exists(output_file):
                os.unlink(output_file)

    def test_find_coi_pdfs_exact_matching(self):
        """Test the find_coi_pdfs function with exact matching."""
        if not self.pdf_dir or not os.path.isdir(self.pdf_dir):
            self.skipTest(f"PDF directory not available: {self.pdf_dir}")
            
        # Get list of PDF files to test with
        pdf_files = [f for f in os.listdir(self.pdf_dir) if f.lower().endswith('.pdf')]
        if not pdf_files:
            self.skipTest("No PDF files found in directory")
            
        # Test with a filename that should match exactly
        test_file = pdf_files[0]
        test_name = os.path.splitext(test_file)[0]  # Remove .pdf extension
        
        # Test exact matching (fuzzy disabled)
        results = find_coi_pdfs(
            pdf_directory_path=self.pdf_dir,
            subcontractor_name=test_name,
            fuzzy_config={'enabled': False}
        )
        
        self.assertIsInstance(results, list, "Should return a list")
        # Should find at least one match for exact filename
        if test_name.replace('_', '').replace('-', '').replace(' ', '').isalnum():
            # Only expect exact match if the name is reasonably simple
            self.assertTrue(len(results) >= 1, f"Should find exact match for '{test_name}'")

    def test_find_coi_pdfs_fuzzy_matching(self):
        """Test the find_coi_pdfs function with fuzzy matching."""
        if not self.pdf_dir or not os.path.isdir(self.pdf_dir):
            self.skipTest(f"PDF directory not available: {self.pdf_dir}")
            
        # Test with a name that might have fuzzy matches
        test_name = "ABC Construction"  # Generic name that might have partial matches
        
        results = find_coi_pdfs(
            pdf_directory_path=self.pdf_dir,
            subcontractor_name=test_name,
            fuzzy_config={'enabled': True, 'threshold': 50.0, 'max_results': 5}
        )
        
        self.assertIsInstance(results, list, "Should return a list")
        # Results may be empty if no fuzzy matches above threshold, which is fine

    @patch('coi_auditor.main.diagnose_pdf_discovery')
    @patch('coi_auditor.main.sys.exit')
    def test_run_diagnostic_subcontractor_function(self, mock_exit, mock_diagnose):
        """Test the run_diagnostic_subcontractor function."""
        # Mock the diagnostic function to return test data
        mock_diagnose.return_value = {
            'subcontractor_name': 'Test Contractor',
            'timestamp': '2024-01-01T00:00:00',
            'summary': {
                'status': 'success_exact',
                'directory_accessible': True,
                'pdf_files_found': 5,
                'exact_matches': 1,
                'fuzzy_matches': 0,
                'rapidfuzz_available': True,
                'name_variations_generated': 2
            },
            'directory_analysis': {
                'configured_path': '/test/path',
                'path_exists': True,
                'is_directory': True,
                'accessible': True,
                'effective_search_dir': '/test/path',
                'directory_type': 'recognized_coi_directory',
                'pdf_count': 5,
                'sample_files': ['test1.pdf', 'test2.pdf']
            },
            'name_analysis': {
                'original_name': 'Test Contractor',
                'simple_normalized': 'testcontractor',
                'enhanced_normalized': 'testcontractor',
                'all_variations': ['testcontractor', 'test contractor'],
                'variation_count': 2
            },
            'exact_matching': {
                'matches_found': 1,
                'matched_files': ['test_contractor.pdf']
            },
            'config_used': {
                'fuzzy_matching_enabled': True,
                'fuzzy_threshold': 75.0,
                'fuzzy_algorithms': ['ratio', 'partial_ratio'],
                'expected_folder_name': 'Subcontractor COIs',
                'alternative_folder_names': []
            },
            'recommendations': ['Test recommendation 1', 'Test recommendation 2']
        }
        
        # Test successful case
        config = {'pdf_directory_path': '/test/path'}
        run_diagnostic_subcontractor(
            subcontractor_name='Test Contractor',
            pdf_directory=None,
            output_file=None,
            config=config
        )
        
        # Verify the diagnostic function was called
        mock_diagnose.assert_called_once_with(
            subcontractor_name='Test Contractor',
            pdf_directory_path='/test/path',
            output_file=None
        )
        
        # Verify sys.exit was called with success code
        mock_exit.assert_called_once_with(0)


def test_fuzzy_matching():
    """Legacy test function for backward compatibility."""
    test_case = TestFuzzyMatching()
    test_case.setUpClass()
    test_case.test_legacy_fuzzy_matching()


if __name__ == '__main__':
    # Run the legacy test for backward compatibility
    test_fuzzy_matching()
    
    # Run all unit tests
    unittest.main(verbosity=2)
