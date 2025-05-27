#!/usr/bin/env python3
"""
Environment validation script for COI Auditor production setup.
This script verifies that all required components are properly configured
and ready for pipeline execution.
"""

import os
import sys
from pathlib import Path
from datetime import datetime
import logging
from typing import List, Tuple, Dict, Any

# Setup basic logging for validation
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class ValidationError(Exception):
    """Custom exception for validation failures."""
    pass

class EnvironmentValidator:
    """Validates the COI Auditor environment configuration."""
    
    def __init__(self):
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.project_root = Path(__file__).resolve().parent.parent
        
    def validate_dotenv_file(self) -> bool:
        """Validate that .env file exists and contains required variables."""
        logger.info("🔍 Validating .env file...")
        
        dotenv_path = self.project_root / '.env'
        if not dotenv_path.exists():
            self.errors.append("❌ .env file not found in project root")
            return False
            
        # Load .env file manually to check contents
        required_vars = [
            'EXCEL_FILE_PATH',
            'PDF_DIRECTORY_PATH', 
            'OUTPUT_DIRECTORY_PATH',
            'AUDIT_START_DATE',
            'AUDIT_END_DATE',
            'EXCEL_HEADER_ROW',
            'GL_FROM_COL',
            'GL_TO_COL',
            'WC_FROM_COL',
            'WC_TO_COL'
        ]
        
        env_content = dotenv_path.read_text(encoding='utf-8')
        missing_vars = []
        
        for var in required_vars:
            if f"{var}=" not in env_content:
                missing_vars.append(var)
                
        if missing_vars:
            self.errors.append(f"❌ Missing required environment variables: {', '.join(missing_vars)}")
            return False
            
        logger.info("✅ .env file validation passed")
        return True
        
    def validate_paths(self) -> bool:
        """Validate that all configured paths exist and are accessible."""
        logger.info("🔍 Validating configured paths...")
        
        # Load environment variables
        from dotenv import load_dotenv
        load_dotenv(self.project_root / '.env')
        
        paths_to_check = {
            'EXCEL_FILE_PATH': ('Excel file', 'file'),
            'PDF_DIRECTORY_PATH': ('PDF directory', 'directory'),
            'OUTPUT_DIRECTORY_PATH': ('Output directory', 'directory')
        }
        
        all_valid = True
        
        for env_var, (description, path_type) in paths_to_check.items():
            path_str = os.getenv(env_var)
            if not path_str:
                self.errors.append(f"❌ {env_var} not set in environment")
                all_valid = False
                continue
                
            path_obj = Path(path_str)
            
            if path_type == 'file':
                if not path_obj.exists():
                    self.errors.append(f"❌ {description} not found: {path_obj}")
                    all_valid = False
                elif not path_obj.is_file():
                    self.errors.append(f"❌ {description} is not a file: {path_obj}")
                    all_valid = False
                else:
                    logger.info(f"✅ {description} exists: {path_obj}")
                    
            elif path_type == 'directory':
                if not path_obj.exists():
                    if env_var == 'OUTPUT_DIRECTORY_PATH':
                        # Output directory can be created
                        try:
                            path_obj.mkdir(parents=True, exist_ok=True)
                            logger.info(f"✅ Created {description}: {path_obj}")
                        except Exception as e:
                            self.errors.append(f"❌ Cannot create {description}: {path_obj} - {e}")
                            all_valid = False
                    else:
                        self.errors.append(f"❌ {description} not found: {path_obj}")
                        all_valid = False
                elif not path_obj.is_dir():
                    self.errors.append(f"❌ {description} is not a directory: {path_obj}")
                    all_valid = False
                else:
                    logger.info(f"✅ {description} exists: {path_obj}")
                    
        return all_valid
        
    def validate_dates(self) -> bool:
        """Validate audit date configuration."""
        logger.info("🔍 Validating audit dates...")
        
        start_date_str = os.getenv('AUDIT_START_DATE')
        end_date_str = os.getenv('AUDIT_END_DATE')
        
        if not start_date_str or not end_date_str:
            self.errors.append("❌ AUDIT_START_DATE and AUDIT_END_DATE must be set")
            return False
            
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            if start_date >= end_date:
                self.errors.append("❌ AUDIT_START_DATE must be before AUDIT_END_DATE")
                return False
                
            logger.info(f"✅ Audit period: {start_date} to {end_date}")
            return True
            
        except ValueError as e:
            self.errors.append(f"❌ Invalid date format (use YYYY-MM-DD): {e}")
            return False
            
    def validate_dependencies(self) -> bool:
        """Validate that required Python packages are installed."""
        logger.info("🔍 Validating Python dependencies...")
        
        required_packages = [
            'openpyxl',
            'pdfplumber',
            'dotenv',  # python-dotenv package imports as 'dotenv'
            'pytesseract',
            'paddleocr',
            'rich',
            'tqdm',
            'pdf2image',
            'cv2'  # opencv-python
        ]
        
        missing_packages = []
        
        for package in required_packages:
            try:
                if package == 'cv2':
                    import cv2
                else:
                    __import__(package)
                logger.info(f"✅ {package} is installed")
            except ImportError:
                missing_packages.append(package)
                
        if missing_packages:
            self.errors.append(f"❌ Missing required packages: {', '.join(missing_packages)}")
            self.errors.append("   Install with: pip install -e .")
            return False
            
        return True
        
    def validate_test_fixtures(self) -> bool:
        """Validate that test fixtures are available and properly structured."""
        logger.info("🔍 Validating test fixtures...")
        
        pdf_dir = Path(os.getenv('PDF_DIRECTORY_PATH', 'tests/fixtures/'))
        
        if not pdf_dir.exists():
            self.errors.append(f"❌ Test fixtures directory not found: {pdf_dir}")
            return False
            
        pdf_files = list(pdf_dir.glob('*.pdf'))
        if not pdf_files:
            self.errors.append(f"❌ No PDF files found in {pdf_dir}")
            return False
            
        logger.info(f"✅ Found {len(pdf_files)} test PDF files:")
        for pdf_file in sorted(pdf_files):
            logger.info(f"   📄 {pdf_file.name}")
            
        return True
        
    def validate_excel_structure(self) -> bool:
        """Validate the Excel file structure."""
        logger.info("🔍 Validating Excel file structure...")
        
        excel_path = Path(os.getenv('EXCEL_FILE_PATH', ''))
        if not excel_path.exists():
            self.errors.append(f"❌ Excel file not found: {excel_path}")
            return False
            
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            if 'SUMMARY' not in wb.sheetnames:
                self.errors.append("❌ Excel file missing required 'SUMMARY' worksheet")
                return False
                
            ws = wb['SUMMARY']
            header_row = int(os.getenv('EXCEL_HEADER_ROW', '6'))
            
            # Check if header row exists
            if ws.max_row < header_row:
                self.errors.append(f"❌ Excel file has insufficient rows (need at least {header_row})")
                return False
                
            # Check for required columns
            headers = [cell.value for cell in ws[header_row]]
            required_headers = ['Name']  # Minimum required
            
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                self.errors.append(f"❌ Excel file missing required headers: {', '.join(missing_headers)}")
                return False
                
            logger.info(f"✅ Excel file structure validated")
            logger.info(f"   📊 Headers: {headers}")
            logger.info(f"   📈 Data rows: {ws.max_row - header_row}")
            
            return True
            
        except Exception as e:
            self.errors.append(f"❌ Error validating Excel file: {e}")
            return False
            
    def validate_configuration_loading(self) -> bool:
        """Test that the application can load configuration successfully."""
        logger.info("🔍 Testing configuration loading...")
        
        try:
            # Import and test config loading
            sys.path.insert(0, str(self.project_root / 'src'))
            from coi_auditor.config import load_config
            
            config = load_config()
            
            # Verify key configuration items
            required_config_keys = [
                'excel_file_path',
                'pdf_directory_path', 
                'output_directory_path',
                'audit_start_date',
                'audit_end_date'
            ]
            
            missing_keys = [key for key in required_config_keys if key not in config]
            if missing_keys:
                self.errors.append(f"❌ Configuration missing keys: {', '.join(missing_keys)}")
                return False
                
            logger.info("✅ Configuration loading successful")
            return True
            
        except Exception as e:
            self.errors.append(f"❌ Configuration loading failed: {e}")
            return False
            
    def run_validation(self) -> bool:
        """Run all validation checks."""
        logger.info("🚀 Starting COI Auditor environment validation...")
        logger.info("=" * 60)
        
        validation_steps = [
            self.validate_dotenv_file,
            self.validate_paths,
            self.validate_dates,
            self.validate_dependencies,
            self.validate_test_fixtures,
            self.validate_excel_structure,
            self.validate_configuration_loading
        ]
        
        all_passed = True
        
        for step in validation_steps:
            try:
                if not step():
                    all_passed = False
            except Exception as e:
                self.errors.append(f"❌ Validation step failed: {step.__name__} - {e}")
                all_passed = False
                
        logger.info("=" * 60)
        
        if all_passed:
            logger.info("🎉 ALL VALIDATION CHECKS PASSED!")
            logger.info("✅ Environment is ready for COI Auditor pipeline execution")
            return True
        else:
            logger.error("❌ VALIDATION FAILED!")
            logger.error("The following issues must be resolved:")
            for error in self.errors:
                logger.error(f"   {error}")
                
            if self.warnings:
                logger.warning("Warnings:")
                for warning in self.warnings:
                    logger.warning(f"   {warning}")
                    
            return False

def main():
    """Main validation entry point."""
    validator = EnvironmentValidator()
    success = validator.run_validation()
    
    if success:
        print("\n🎯 READY FOR PIPELINE EXECUTION")
        print("You can now run the COI auditor with:")
        print("   python -m coi_auditor.main")
        print("   # or for batch validation:")
        print("   python -m coi_auditor.main --validate-batch tests/fixtures/")
        sys.exit(0)
    else:
        print("\n❌ ENVIRONMENT NOT READY")
        print("Please fix the issues above before running the pipeline.")
        sys.exit(1)

if __name__ == "__main__":
    main()