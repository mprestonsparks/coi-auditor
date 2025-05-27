from pathlib import Path
import openpyxl
import logging

logger = logging.getLogger(__name__)

def workbook_is_clean(path: Path) -> bool:
    """
    Checks if an Excel workbook's "ERROR_REPORT" sheet is clean.

    A workbook is considered clean if its "ERROR_REPORT" worksheet has no rows
    after the header row.

    Args:
        path: The Path to the Excel workbook.

    Returns:
        True if the "ERROR_REPORT" sheet is clean, False otherwise.
    """
    workbook = None  # Initialize to None
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "ERROR_REPORT" not in workbook.sheetnames:
            # If the sheet doesn't exist, we can consider it "clean" for this check,
            # or handle as an error. For now, let's assume no error report means no errors.
            # Alternatively, this could indicate a problem with workbook generation.
            # For the purpose of this validation, if the sheet is missing,
            # it implies no errors were reported in that specific format.
            return True

        error_sheet = workbook["ERROR_REPORT"]

        # Check if there are any rows beyond the header (min_row=2)
        # error_sheet.max_row gives the total number of rows.
        # If max_row is 1 (only header) or 0 (empty sheet), it's clean.
        # Iterating and checking for any row is more robust if sheet might be empty or just header.
        for row_idx, row in enumerate(error_sheet.iter_rows(min_row=2)):
            if any(cell.value is not None for cell in row):
                # Found a row with data after the header
                return False
        return True
    except Exception as e:
        logger.error(f"Error verifying workbook '{path}': {e}")
        # If there's any issue opening or reading the workbook,
        # consider it not clean.
        return False
    finally:
        if workbook:
            workbook.close() # Ensure close is called if workbook was opened

if __name__ == '__main__':
    # Example usage (for testing this module directly)
    # Create dummy excel files for testing
    from openpyxl import Workbook

    # Test case 1: Clean workbook
    wb_clean = Workbook()
    # Remove default sheet if it exists
    if "Sheet" in wb_clean.sheetnames:
        wb_clean.remove(wb_clean["Sheet"])
    error_sheet_clean = wb_clean.create_sheet("ERROR_REPORT")
    error_sheet_clean.append(["Error Type", "Details", "Source Cell"]) # Header
    clean_path = Path("temp_clean_workbook.xlsx")
    wb_clean.save(clean_path)
    print(f"'{clean_path}' is clean: {workbook_is_clean(clean_path)}")
    clean_path.unlink() # Clean up

    # Test case 2: Workbook with errors
    wb_errors = Workbook()
    if "Sheet" in wb_errors.sheetnames:
        wb_errors.remove(wb_errors["Sheet"])
    error_sheet_errors = wb_errors.create_sheet("ERROR_REPORT")
    error_sheet_errors.append(["Error Type", "Details", "Source Cell"]) # Header
    error_sheet_errors.append(["Date Mismatch", "Expected 2023-01-01, got 2023-01-02", "A5"])
    error_sheet_errors.append(["Missing Info", "Insured name not found", "B10"])
    errors_path = Path("temp_errors_workbook.xlsx")
    wb_errors.save(errors_path)
    print(f"'{errors_path}' is clean: {workbook_is_clean(errors_path)}")
    errors_path.unlink() # Clean up

    # Test case 3: Workbook without ERROR_REPORT sheet
    wb_no_report = Workbook()
    if "Sheet" in wb_no_report.sheetnames:
        wb_no_report.remove(wb_no_report["Sheet"])
    no_report_path = Path("temp_no_report_workbook.xlsx")
    wb_no_report.save(no_report_path)
    print(f"'{no_report_path}' (no ERROR_REPORT) is clean: {workbook_is_clean(no_report_path)}")
    no_report_path.unlink() # Clean up

    # Test case 4: Empty ERROR_REPORT sheet
    wb_empty_report = Workbook()
    if "Sheet" in wb_empty_report.sheetnames:
        wb_empty_report.remove(wb_empty_report["Sheet"])
    wb_empty_report.create_sheet("ERROR_REPORT") # Empty error report
    empty_report_path = Path("temp_empty_report_workbook.xlsx")
    wb_empty_report.save(empty_report_path)
    print(f"'{empty_report_path}' (empty ERROR_REPORT) is clean: {workbook_is_clean(empty_report_path)}")
    empty_report_path.unlink() # Clean up