"""Main entry point for the COI Auditor application."""

import logging
import sys
import time
import argparse
from pathlib import Path
from typing import Optional
import tempfile
import platform # Added for OS detection
import shutil # For cleanup, though TemporaryDirectory handles it.
import os # Added missing import
import openpyxl # For reading error reports
from tqdm import tqdm
from rich.console import Console
from rich.theme import Theme
from .terminal_formatter import get_formatter

# Import setup_logging and load_config from .config
from .config import setup_logging, load_config 
from .excel_handler import (
    read_subcontractors,
    update_excel,
    write_gaps_report,
    run_excel_sanity_checks,
    create_error_report_workbook_for_validation
)
from .audit import process_subcontractor
from .verify import workbook_is_clean
from .pdf_parser import extract_raw_ocr_text_from_pdf, extract_dates_from_pdf, diagnose_pdf_discovery # Added for diagnostic and debug modes

# Setup logging as the first step
# Determine project root for default log file path
project_root_for_log_setup = Path(__file__).resolve().parent.parent # src/coi_auditor -> src -> project_root
logs_dir_path_main = project_root_for_log_setup / 'logs'
logs_dir_path_main.mkdir(parents=True, exist_ok=True)
default_log_file = logs_dir_path_main / 'coi_audit_main.log'

# Call setup_logging here. It configures the root logger.
# Subsequent getLogger calls will inherit this configuration.
setup_logging(level="INFO", log_file=str(default_log_file))

# Now, get the logger for this module, which will use the RichHandler.
logger = logging.getLogger(__name__)


# Headers for the ERROR_REPORT sheet in validation mode
VALIDATION_ERROR_REPORT_HEADERS = ["Subcontractor Name", "Subcontractor ID", "Issue Type", "Details", "Policy Type", "Effective Date", "Expiration Date", "File Path", "Page Number"]


def run_diagnostic_pdf_ocr(pdf_path_str: str, config: dict):
    """
    Performs OCR on a single PDF and prints/saves the raw text.
    """
    pdf_path = Path(pdf_path_str)
    if not pdf_path.is_file():
        logger.error(f"[bold red]Diagnostic PDF not found: [yellow]{pdf_path}[/yellow][/bold red]")
        print(f"Error: Diagnostic PDF '{pdf_path}' does not exist.")
        sys.exit(1)

    logger.info(f"Starting diagnostic OCR for PDF: [cyan]{pdf_path}[/cyan]")
    
    notes = []  # Initialize notes list for the function
    raw_text = extract_raw_ocr_text_from_pdf(pdf_path, notes)

    print("\n--- RAW OCR TEXT ---")
    print(raw_text)
    print("--- END RAW OCR TEXT ---\n")

    output_dir_path = Path(config.get('output_directory_path', '.'))
    output_dir_path.mkdir(parents=True, exist_ok=True)
    output_filename = f"diagnostic_ocr_output_{pdf_path.stem}.txt"
    output_file_path = output_dir_path / output_filename
    
    try:
        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write(f"Diagnostic OCR Output for: {pdf_path.name}\n")
            f.write("="*30 + "\n")
            f.write(raw_text)
        logger.info(f"Diagnostic OCR text saved to: [green]{output_file_path}[/green]")
        print(f"Diagnostic OCR text also saved to: {output_file_path}")
    except IOError as e:
        logger.error(f"[bold red]Failed to save diagnostic OCR text to {output_file_path}: {e}[/bold red]")
        print(f"Error: Could not save diagnostic OCR text to {output_file_path}: {e}")

    logger.info("Diagnostic OCR completed. Exiting.")
    sys.exit(0)


def run_debug_single_pdf(pdf_path_str: str, config: dict):
    """
    Processes a single PDF in debug mode, printing date tokens and table snippets.
    """
    pdf_path = Path(pdf_path_str)
    if not pdf_path.is_file():
        logger.error(f"[bold red]Debug PDF not found: [yellow]{pdf_path}[/yellow][/bold red]")
        print(f"Error: Debug PDF '{pdf_path}' does not exist.")
        sys.exit(1)

    logger.info(f"Starting debug processing for PDF: [cyan]{pdf_path}[/cyan]")
    
    # Set a flag in config or pass a parameter to indicate debug mode
    # For now, we'll rely on a new parameter to extract_dates_from_pdf
    
    # This call will need to be to a modified extract_dates_from_pdf or a wrapper
    # that handles the debug printing.
    # For now, let's assume extract_dates_from_pdf will be modified to accept a debug flag.
    
    print(f"\n--- DEBUG MODE: Processing {pdf_path.name} ---")
    
    # The actual debug prints for tokens and snippets will happen within pdf_parser.py
    # We just invoke the main extraction function here with the debug context.
    # The `extract_dates_from_pdf` function returns (dates_dict, notes_list)
    # We might not need to do much with the results here, as the goal is the printed debug output.
    
    _extracted_dates, _notes = extract_dates_from_pdf(pdf_path) # Extract dates for debug analysis
    
    # The notes might already contain some of the debug info if we choose to log it there too.
    # For direct console output of tokens/snippets, pdf_parser.py will handle it.

    print(f"--- DEBUG MODE: Finished processing {pdf_path.name} ---")
    logger.info(f"Debug processing for PDF: [cyan]{pdf_path}[/cyan] completed. Exiting.")
    sys.exit(0)


def run_diagnostic_subcontractor(subcontractor_name: str, pdf_directory: Optional[str] = None, output_file: Optional[str] = None, config: Optional[dict] = None):
    """
    Runs diagnostic analysis for a specific subcontractor's PDF discovery.
    """
    logger.info(f"Starting diagnostic analysis for subcontractor: '{subcontractor_name}'")
    
    # Use config PDF directory if not provided
    effective_pdf_dir = pdf_directory or config.get('pdf_directory_path', '') if config else ''
    
    if not effective_pdf_dir:
        logger.error("No PDF directory specified and none found in configuration")
        print("Error: No PDF directory specified. Use --pdf-directory or configure PDF_DIRECTORY_PATH in .env")
        sys.exit(1)
    
    # Run the diagnostic
    results = diagnose_pdf_discovery(
        subcontractor_name=subcontractor_name,
        pdf_directory_path=effective_pdf_dir,
        output_file=output_file
    )
    
    # Display results to console
    print(f"\n=== PDF Discovery Diagnostic Results for '{subcontractor_name}' ===")
    print(f"Timestamp: {results['timestamp']}")
    print(f"Status: {results['summary']['status']}")
    
    # Directory Analysis
    print(f"\n--- Directory Analysis ---")
    dir_analysis = results['directory_analysis']
    print(f"Configured Path: {dir_analysis['configured_path']}")
    print(f"Path Exists: {dir_analysis['path_exists']}")
    print(f"Is Directory: {dir_analysis['is_directory']}")
    print(f"Accessible: {dir_analysis['accessible']}")
    if dir_analysis['accessible']:
        print(f"Effective Search Directory: {dir_analysis.get('effective_search_dir', 'N/A')}")
        print(f"Directory Type: {dir_analysis.get('directory_type', 'N/A')}")
        print(f"PDF Files Found: {dir_analysis['pdf_count']}")
        if dir_analysis['sample_files']:
            print(f"Sample Files: {', '.join(dir_analysis['sample_files'][:5])}")
            if len(dir_analysis['sample_files']) > 5:
                print(f"  ... and {len(dir_analysis['sample_files']) - 5} more")
    
    # Name Analysis
    print(f"\n--- Name Analysis ---")
    name_analysis = results['name_analysis']
    print(f"Original Name: '{name_analysis['original_name']}'")
    print(f"Simple Normalized: '{name_analysis['simple_normalized']}'")
    print(f"Enhanced Normalized: '{name_analysis['enhanced_normalized']}'")
    print(f"Generated Variations ({name_analysis['variation_count']}): {name_analysis['all_variations']}")
    
    # Matching Results
    if 'exact_matching' in results:
        print(f"\n--- Exact Matching Results ---")
        exact = results['exact_matching']
        print(f"Matches Found: {exact['matches_found']}")
        if exact['matched_files']:
            print(f"Matched Files: {', '.join(exact['matched_files'])}")
    
    if 'fuzzy_matching' in results:
        print(f"\n--- Fuzzy Matching Results ---")
        fuzzy = results['fuzzy_matching']
        print(f"RapidFuzz Available: {fuzzy['rapidfuzz_available']}")
        print(f"Threshold Used: {fuzzy['threshold_used']}%")
        print(f"Matches Found: {fuzzy['matches_found']}")
        if fuzzy['matched_files']:
            print(f"Matched Files: {', '.join(fuzzy['matched_files'])}")
        
        if fuzzy['all_scores']:
            print(f"\nTop Fuzzy Match Scores:")
            for i, (filename, score) in enumerate(fuzzy['all_scores'][:10], 1):
                print(f"  {i:2d}. {filename:<40} {score:6.1f}%")
    
    # Configuration
    print(f"\n--- Configuration Used ---")
    config_used = results['config_used']
    print(f"Fuzzy Matching Enabled: {config_used['fuzzy_matching_enabled']}")
    print(f"Fuzzy Threshold: {config_used['fuzzy_threshold']}%")
    print(f"Fuzzy Algorithms: {', '.join(config_used['fuzzy_algorithms'])}")
    print(f"Expected COI Folder: '{config_used['expected_folder_name']}'")
    if config_used['alternative_folder_names']:
        print(f"Alternative Folders: {', '.join(config_used['alternative_folder_names'])}")
    
    # Recommendations
    print(f"\n--- Recommendations ---")
    for i, recommendation in enumerate(results['recommendations'], 1):
        print(f"  {i}. {recommendation}")
    
    # Summary
    print(f"\n--- Summary ---")
    summary = results['summary']
    print(f"Directory Accessible: {summary['directory_accessible']}")
    print(f"PDF Files Found: {summary['pdf_files_found']}")
    print(f"Exact Matches: {summary['exact_matches']}")
    print(f"Fuzzy Matches: {summary['fuzzy_matches']}")
    print(f"Name Variations Generated: {summary['name_variations_generated']}")
    print(f"RapidFuzz Available: {summary['rapidfuzz_available']}")
    
    # Output file info
    if 'output_file_saved' in results:
        print(f"\nDetailed results saved to: {results['output_file_saved']}")
    elif 'output_file_error' in results:
        print(f"\nError saving output file: {results['output_file_error']}")
    
    print(f"\n=== Diagnostic Complete ===")
    
    # Exit with appropriate code
    if summary['status'] in ['success_exact', 'success_fuzzy']:
        logger.info(f"Diagnostic completed successfully for '{subcontractor_name}'")
        sys.exit(0)
    else:
        logger.warning(f"Diagnostic found issues for '{subcontractor_name}'. See recommendations above.")
        sys.exit(1)


def perform_single_pdf_audit(pdf_path: Path, output_excel_path: Path, config: dict):
    """
    Processes a single PDF, generates an audit report (specifically an ERROR_REPORT sheet),
    and saves it to the specified Excel path.
    """
    logger.info(f"Auditing PDF for validation: [cyan]{pdf_path}[/cyan]")
    
    dummy_sub = {
        'name': pdf_path.stem,
        'id': f"fixture_{pdf_path.stem}", 
        'coi_pdf_path': str(pdf_path),
        'row': 1, 
        'expected_insured_name': None,
        'gl_policy_num_expected': None,
        'wc_policy_num_expected': None,
    }

    try:
        sub_result, sub_gaps = process_subcontractor(dummy_sub, config, direct_pdf_path=pdf_path)
        create_error_report_workbook_for_validation(output_excel_path, sub_gaps, VALIDATION_ERROR_REPORT_HEADERS)
        logger.info(f"Temporary audit report for [cyan]{pdf_path.name}[/cyan] saved to [green]{output_excel_path}[/green]")

    except Exception as e:
        logger.error(f"[bold red]Failed during single PDF audit for [cyan]{pdf_path.name}[/cyan]: {e}[/bold red]", exc_info=True)
        error_gap = [{
            'Subcontractor Name': pdf_path.stem,
            'Subcontractor ID': f"fixture_{pdf_path.stem}",
            'Issue Type': "Audit Processing Error",
            'Details': f"Critical error during audit: {e}",
            'Policy Type': "N/A",
            'Effective Date': "N/A",
            'Expiration Date': "N/A",
            'File Path': str(pdf_path),
            'Page Number': "N/A"
        }]
        create_error_report_workbook_for_validation(output_excel_path, error_gap, VALIDATION_ERROR_REPORT_HEADERS)
        logger.info(f"Error audit report for [cyan]{pdf_path.name}[/cyan] saved to [green]{output_excel_path}[/green]")


def run_batch_validation(fixtures_dir_str: str):
    """
    Runs the audit process for all PDFs in a fixtures directory and verifies
    that their "ERROR_REPORT" sheets are clean.
    """
    formatter = get_formatter()
    formatter.print_header("COI AUDITOR - BATCH VALIDATION MODE")
    
    logger.info(f"Starting batch validation for fixtures in: {fixtures_dir_str}")
    config = load_config()  # Validation mode uses production config by default

    fixtures_path = Path(fixtures_dir_str)
    if not fixtures_path.is_dir():
        formatter.error(f"Fixtures directory not found: {fixtures_path}")
        sys.exit(1)

    pdf_files_to_process = sorted(list(fixtures_path.glob('*.pdf')))

    if not pdf_files_to_process:
        formatter.warning(f"No PDF files found in fixtures directory: {fixtures_path}")
        sys.exit(0)
    
    formatter.info(f"Processing {len(pdf_files_to_process)} PDF file(s) from {fixtures_path}")

    overall_success = True
    failed_reports_details = [] 

    cleanup_args = {}
    if platform.system() == "Windows":
        logger.info("Windows OS detected, setting ignore_cleanup_errors=True for TemporaryDirectory.")
        cleanup_args['ignore_cleanup_errors'] = True

    with tempfile.TemporaryDirectory(prefix="coi_val_", **cleanup_args) as tmpdir:
        tmpdir_path = Path(tmpdir)
        logger.info(f"Using temporary directory for reports: {tmpdir_path}")

        # Start progress display
        formatter.start_progress(len(pdf_files_to_process), "Validating PDF Fixtures")
        
        try:
            for i, pdf_file in enumerate(pdf_files_to_process):
                formatter.update_progress(i, pdf_file.name, 'Processing')
                logger.info(f"--- Processing fixture: {pdf_file.name} ---")
                temp_excel_name = f"{pdf_file.stem}_audit_report.xlsx"
                temp_excel_path = tmpdir_path / temp_excel_name
                
                try:
                    perform_single_pdf_audit(pdf_file, temp_excel_path, config)
                    
                    if not temp_excel_path.exists():
                        logger.error(f"Audit report was not created for {pdf_file.name} at {temp_excel_path}")
                        overall_success = False
                        failed_reports_details.append((pdf_file.name, [f"| Critical Error | Report file not created"]))
                        formatter.update_progress(i + 1, pdf_file.name, 'ERROR')
                        continue

                    if not workbook_is_clean(temp_excel_path):
                        overall_success = False
                        logger.warning(f"Validation FAILED for {pdf_file.name}. Report: {temp_excel_path}")
                        formatter.update_progress(i + 1, pdf_file.name, 'FAILED')
                        current_pdf_errors = []
                        try:
                            wb = openpyxl.load_workbook(temp_excel_path, read_only=True, data_only=True)
                            if "ERROR_REPORT" in wb.sheetnames:
                                error_sheet = wb["ERROR_REPORT"]
                                has_errors = False
                                for row_idx, row_values in enumerate(error_sheet.iter_rows(min_row=2, values_only=True)):
                                    if any(cell is not None for cell in row_values):
                                        has_errors = True
                                        error_details = " | ".join(str(cell).strip() if cell is not None else "" for cell in row_values)
                                        current_pdf_errors.append(error_details)
                                if not has_errors and error_sheet.max_row > 1 :
                                    logger.info(f"ERROR_REPORT for {pdf_file.name} has rows but they appear empty after header.")
                                elif not has_errors:
                                    logger.warning(f"workbook_is_clean was False for {pdf_file.name}, but no error rows extracted from ERROR_REPORT.")
                                    current_pdf_errors.append("| Internal Discrepancy | workbook_is_clean reported errors, but none found in sheet.")
                            else:
                                logger.error(f"ERROR_REPORT sheet missing in {temp_excel_path}, but workbook_is_clean indicated errors.")
                                current_pdf_errors.append("| Critical Error | ERROR_REPORT sheet missing, inconsistent with clean check")
                            
                            if not current_pdf_errors:
                                 current_pdf_errors.append("| Unknown Error | Workbook deemed not clean, but no specific errors extracted.")
                            failed_reports_details.append((pdf_file.name, current_pdf_errors))

                        except Exception as e_read:
                            logger.error(f"Error reading error details from {temp_excel_path} for {pdf_file.name}: {e_read}")
                            failed_reports_details.append((pdf_file.name, [f"| Read Error | Could not parse report: {e_read}"]))
                    else:
                        logger.info(f"Validation PASSED for {pdf_file.name}.")
                        formatter.update_progress(i + 1, pdf_file.name, 'OK')

                except Exception as e_process:
                    logger.error(f"Critical error processing {pdf_file.name} during validation: {e_process}", exc_info=True)
                    overall_success = False
                    failed_reports_details.append((pdf_file.name, [f"| Processing Crash | Audit function failed: {e_process}"]))
                    formatter.update_progress(i + 1, pdf_file.name, 'ERROR')
        finally:
            # Always stop progress display
            formatter.stop_progress()

    if overall_success:
        formatter.print_header("BATCH VALIDATION COMPLETED SUCCESSFULLY")
        formatter.success("All PDF fixtures processed successfully and are clean.")
        sys.exit(0)
    else:
        formatter.print_header("BATCH VALIDATION FAILED")
        formatter.error("Batch validation FAILED. Errors found in one or more COIs.")
        
        print("\nDetailed Error Report:")
        print("=" * 80)
        for pdf_name, error_list in failed_reports_details:
            if error_list:
                for error_detail_str in error_list:
                    formatted_line = formatter.colorize(pdf_name, 'yellow') + f" | {error_detail_str}"
                    print(formatted_line)
            else:
                formatted_line = formatter.colorize(pdf_name, 'yellow') + " | An unspecified error occurred or error report was problematic."
                print(formatted_line)
        
        print("=" * 80)
        formatter.error("Failure summary: One or more COIs had errors reported or failed processing.")
        sys.exit(1)


def run_audit():
    """Executes the full COI audit process based on the main Excel file."""
    start_time = time.time()
    formatter = get_formatter()
    
    # Print a nice header
    formatter.print_header("COI AUDITOR - CERTIFICATE OF INSURANCE AUDIT SYSTEM")
    
    logger.info("Starting COI Audit Process...")

    # 1. Load Configuration
    formatter.print_section("Configuration Loading")
    logger.info("Loading configuration...")
    try:
        config = load_config()  # run_audit() uses production config by default
        logger.debug(f"Configuration loaded: {config}")
        formatter.success("Configuration loaded successfully")
    except ValueError as ve:
        logger.error(f"Configuration Error: {ve}", exc_info=True)
        formatter.error(f"Configuration failed: {ve}")
        _write_fatal_error(ve)
        sys.exit(1)

    # 2. Run sanity checks before anything else
    formatter.print_section("Excel File Validation")
    logger.info("Running Excel sanity checks...")
    try:
        run_excel_sanity_checks(config['excel_file_path'])
        formatter.success("Excel file validation passed")
    except Exception as sanity_e:
        logger.error(f"Excel sanity check failed: {sanity_e}", exc_info=True)
        formatter.error(f"Excel validation failed: {sanity_e}")
        _write_fatal_error(sanity_e)
        sys.exit(1)

    try:
        # 3. Read Subcontractor List
        formatter.print_section("Subcontractor Processing")
        logger.info(f"Reading subcontractors from Excel: {config['excel_file_path']}")
        subcontractors, headers = read_subcontractors(config['excel_file_path'])
        if not subcontractors:
            formatter.warning("No subcontractors found in the Excel file. Exiting.")
            return

        # 3. Process Each Subcontractor
        all_results = []
        all_gaps = []
        total_subs = len(subcontractors)
        logger.info(f"Found {total_subs} subcontractors to process.")

        # Start our custom progress display
        formatter.start_progress(total_subs, "Auditing Subcontractors")
        
        try:
            for i, sub in enumerate(subcontractors):
                sub_name = sub.get('name', 'N/A')
                
                # Update progress display
                formatter.update_progress(i, sub_name[:40], 'Processing')
                
                logger.info(f"Starting audit for subcontractor: {sub_name} (Row: {sub.get('row', 'N/A')})")
                try:
                    sub_result, sub_gaps = process_subcontractor(sub, config)
                    all_results.append(sub_result)
                    all_gaps.extend(sub_gaps)
                    
                    # Determine status based on gaps
                    status = 'ERROR' if sub_gaps else 'OK'
                    formatter.update_progress(i + 1, sub_name[:40], status)
                    
                except Exception as sub_proc_e:
                    logger.error(f"ERROR processing subcontractor {sub.get('name', 'N/A')} (Row: {sub.get('row', 'N/A')}): {sub_proc_e}", exc_info=True)
                    all_results.append({
                        'row': sub.get('row'),
                        'name': sub.get('name', 'N/A'),
                        'id': sub.get('id', 'N/A'),
                        'notes': f"ERROR PROCESSING: {sub_proc_e}",
                        'gl_gap_status': 'Error',
                        'wc_gap_status': 'Error'
                    })
                    all_gaps.append({
                         'subcontractor_name': sub.get('name', 'N/A'),
                         'subcontractor_id': sub.get('id', 'N/A'),
                         'issue_type': 'Processing Error',
                         'details': str(sub_proc_e)
                    })
                    formatter.update_progress(i + 1, sub_name[:40], 'ERROR')
        finally:
            # Always stop progress display
            formatter.stop_progress()

        formatter.success("Finished processing all subcontractors")

        # 4. Update Excel File
        formatter.print_section("Excel File Update")
        logger.info(f"Updating Excel file: {config['excel_file_path']}")
        update_excel(config['excel_file_path'], all_results, headers)
        formatter.success("Excel file updated successfully")

        # 5. Write Gaps Report (CSV and Excel worksheet)
        formatter.print_section("Gaps Report Generation")
        logger.info(f"Writing gaps report to: {config['output_directory_path']} and GAPS_REPORT worksheet in Excel")
        write_gaps_report(config['output_directory_path'], all_gaps, excel_path_str=config['excel_file_path'])
        formatter.success("Gaps report generated successfully")

        end_time = time.time()
        
        # Print completion summary
        formatter.print_header("AUDIT COMPLETED SUCCESSFULLY")
        formatter.success(f"COI Audit Process completed in {end_time - start_time:.2f} seconds")
        formatter.info(f"Processed {total_subs} subcontractors")
        formatter.info(f"Found {len(all_gaps)} total issues")
        
        if all_gaps:
            formatter.warning(f"Review the gaps report for {len(all_gaps)} issues that need attention")
        else:
            formatter.success("No coverage gaps detected - all subcontractors have valid COI coverage")

    except FileNotFoundError as fnf_e:
        formatter.error(f"File Not Found Error: {fnf_e}. Please check paths in .env file.")
        _write_fatal_error(fnf_e)
    except ValueError as val_e:
        formatter.error(f"Configuration or Data Error: {val_e}. Please check .env file or input Excel file.")
        _write_fatal_error(val_e)
    except PermissionError as perm_e:
        formatter.error(f"Permission Error: {perm_e}. Ensure files are not open and permissions are correct.")
        _write_fatal_error(perm_e)
    except ImportError as imp_e:
        formatter.error(f"Import Error: {imp_e}. Ensure all dependencies are installed (pip install -r requirements.txt). Run as 'python -m coi_auditor.main'")
        _write_fatal_error(imp_e)
    except Exception as e:
        formatter.error(f"An unexpected error occurred during the audit process: {e}")
        _write_fatal_error(e)

def _write_fatal_error(exc):
    import traceback
    # Use the consistently defined logs_dir_path from the logging setup
    err_path = default_log_file.parent / 'fatal_error.txt' # Use default_log_file's parent
    try:
        with open(str(err_path), 'w', encoding='utf-8') as f: # open expects string path
            f.write('FATAL ERROR DURING AUDIT\n')
            f.write(traceback.format_exc())
        logger.info(f"Fatal error details written to: [cyan]{err_path}[/cyan]")
    except Exception as file_exc:
        # Use logger here if possible, though print is a fallback if logging itself failed
        logger.error(f"[bold red]Could not write fatal error file to {err_path}: {file_exc}[/bold red]")
        print(f"CRITICAL: Could not write fatal error file to {err_path}: {file_exc}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="COI Auditor: Process COI PDFs and validate.")
    parser.add_argument(
        "--validate-batch",
        metavar="FIXTURES_DIRECTORY",
        type=str,
        help="Run validation harness on all PDFs in the specified directory."
    )
    parser.add_argument(
        "--diagnose-pdf",
        metavar="PDF_FILE_PATH",
        type=str,
        help="Run diagnostic OCR on a single PDF and output the raw text."
    )
    parser.add_argument(
        "--debug-single-pdf",
        metavar="PDF_FILE_PATH",
        type=str,
        help="Process a single PDF with detailed debug output for date tokens and table snippets."
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Run in test mode using test fixtures instead of production data."
    )
    parser.add_argument(
        "--diagnose",
        metavar="SUBCONTRACTOR_NAME",
        type=str,
        help="Run diagnostic analysis for PDF discovery issues with the specified subcontractor name."
    )
    parser.add_argument(
        "--pdf-directory",
        metavar="PDF_DIRECTORY_PATH",
        type=str,
        help="PDF directory path for diagnostic mode (optional, uses config default if not specified)."
    )
    parser.add_argument(
        "--output-file",
        metavar="OUTPUT_FILE_PATH",
        type=str,
        help="Output file path for detailed diagnostic results in JSON format (optional)."
    )
    
    args = parser.parse_args()

    config = load_config(test_mode=args.test) # Load config early for all modes, pass test_mode flag

    if args.debug_single_pdf:
        logger.info(f"Debug mode triggered for PDF: [cyan]{args.debug_single_pdf}[/cyan]")
        run_debug_single_pdf(args.debug_single_pdf, config)
        # run_debug_single_pdf will sys.exit()

    if args.diagnose_pdf:
        logger.info(f"Diagnostic mode triggered for PDF: [cyan]{args.diagnose_pdf}[/cyan]")
        run_diagnostic_pdf_ocr(args.diagnose_pdf, config)
        # run_diagnostic_pdf_ocr will sys.exit()

    if args.diagnose:
        logger.info(f"Diagnostic mode triggered for subcontractor: [cyan]{args.diagnose}[/cyan]")
        run_diagnostic_subcontractor(
            subcontractor_name=args.diagnose,
            pdf_directory=args.pdf_directory,
            output_file=args.output_file,
            config=config
        )
        # run_diagnostic_subcontractor will sys.exit()

    if args.validate_batch:
        os.environ['COI_VALIDATION_MODE'] = '1'
        logger.info("COI_VALIDATION_MODE environment variable set due to --validate-batch flag.")


    if args.validate_batch:
        fixtures_target_path = Path("tests/fixtures/")
        if not fixtures_target_path.exists():
            logger.warning(f"[yellow]Recommended fixtures directory '{fixtures_target_path}' does not exist. Please create it and add test PDFs.[/yellow]")
        elif not any(fixtures_target_path.glob('*.pdf')) and Path(args.validate_batch) == fixtures_target_path:
             logger.warning(f"[yellow]Fixtures directory '{fixtures_target_path}' is empty. Add PDFs for validation.[/yellow]")
        
        run_batch_validation(args.validate_batch)
    elif not args.diagnose_pdf and not args.debug_single_pdf and not args.diagnose: # Only run full audit if not in other modes
        run_audit()
