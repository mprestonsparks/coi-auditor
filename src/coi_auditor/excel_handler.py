"""Handles reading and writing data to the Excel subcontractor list."""

import openpyxl
import csv
import os
from datetime import date
import datetime as dt # Alias for clarity if datetime.datetime is used
import logging
# import gc # Removed, typically not needed
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional # Added for type hints
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet # For type hinting sheet parameters
from .audit import STATUS_GAP # Import necessary constants

logger = logging.getLogger(__name__)

# Define expected/output column headers
# Input columns (expected, but now configurable via .env)
SUBCONTRACTOR_ID_COL = 'Subcontractor ID'  # Optional, used if present

# Output columns (will be created/updated) - These match the Excel column headers
# These seem to be keys for col_map rather than actual header names now.
GL_FROM_KEY = 'GL_FROM_COL'
GL_TO_KEY = 'GL_TO_COL'
WC_FROM_KEY = 'WC_FROM_COL'
WC_TO_KEY = 'WC_TO_COL'

# Actual header names for gap/notes columns if created in a detailed report (not SUMMARY)
GL_GAP_HEADER = 'GL Coverage Gap'
WC_GAP_HEADER = 'WC Coverage Gap'
NOTES_HEADER = 'Audit Notes'

# This mapping seems to be internal to update_excel now.
# GL_OUTPUT_COLUMNS = [GL_FROM_KEY, GL_TO_KEY]
# WC_OUTPUT_COLUMNS = [WC_FROM_KEY, WC_TO_KEY]
# GAP_OUTPUT_COLUMNS = [GL_GAP_HEADER, WC_GAP_HEADER, NOTES_HEADER]


def _find_last_nonempty_row(sheet: Worksheet, col_idx: int, start_row: int, max_row: int) -> int:
    """Find last non-empty row in the specified column (from bottom up)."""
    for row in range(max_row, start_row - 1, -1):
        val = sheet.cell(row=row, column=col_idx).value
        if val and str(val).strip():
            return row
    return start_row - 1  # If all empty, return before start


def read_subcontractors(excel_path_str: str) -> Tuple[List[Dict[str, Any]], List[Any]]:
    """Reads subcontractor data from the specified Excel file, specifically from the 'SUMMARY' worksheet."""
    import traceback # Keep local import for this specific error formatting
    
    excel_path = Path(excel_path_str)
    subcontractors: List[Dict[str, Any]] = []
    headers: List[Any] = []

    try:
        logger.info(f"[read_subcontractors] Attempting to read Excel file: {excel_path}")
        # Read environment variables here, after .env is loaded
        header_row = int(os.getenv('EXCEL_HEADER_ROW', '6'))
        subcontractor_name_col_header = os.getenv('EXCEL_SUBCONTRACTOR_NAME_COL', 'Name')
        # SUBCONTRACTOR_ID_COL is globally defined
        logger.info(f"[read_subcontractors] Config: header_row={header_row}, name_col_header='{subcontractor_name_col_header}', id_col_header='{SUBCONTRACTOR_ID_COL}'")

        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        if 'SUMMARY' not in workbook.sheetnames:
            raise ValueError("Required worksheet 'SUMMARY' not found in workbook.")
        sheet: Worksheet = workbook['SUMMARY']
        logger.info(f"[read_subcontractors] Loaded 'SUMMARY' worksheet. Max row: {sheet.max_row}")
        
        current_headers = [cell.value for cell in sheet[header_row]]
        logger.debug(f"Excel headers found (row {header_row}): {current_headers}")
        headers = current_headers # Store for return

        if subcontractor_name_col_header not in headers:
            raise ValueError(f"Missing required column '{subcontractor_name_col_header}' in Excel file: {excel_path}")

        name_col_idx = headers.index(subcontractor_name_col_header)
        id_col_idx = headers.index(SUBCONTRACTOR_ID_COL) if SUBCONTRACTOR_ID_COL in headers else None
        logger.info(f"[read_subcontractors] Name col idx (0-based): {name_col_idx}, ID col idx (0-based): {id_col_idx}")

        flag_col_env = os.getenv('EXCEL_SUBCONTRACTOR_FLAG_COL')
        flag_col_idx: Optional[int] = None
        if flag_col_env and flag_col_env.strip():
            try:
                if flag_col_env.strip().isdigit():
                    flag_col_idx = int(flag_col_env.strip()) - 1
                else:
                    flag_col_idx = column_index_from_string(flag_col_env.strip().upper()) - 1
            except Exception as e_flag_col:
                logger.warning(f"Invalid EXCEL_SUBCONTRACTOR_FLAG_COL value '{flag_col_env}': {e_flag_col}. Falling back to auto-detection.")
        
        if flag_col_idx is None: # Auto-detect
            for idx, hdr_val in enumerate(headers):
                if hdr_val and str(hdr_val).strip().lower() in ("subcontractor", "subcontractor?", "is subcontractor", "sub?", "sub flag"):
                    flag_col_idx = idx
                    break
        if flag_col_idx is None: flag_col_idx = 2 # Final fallback to column C (index 2)
        logger.info(f"[read_subcontractors] Subcontractor flag col idx (0-based): {flag_col_idx}")

        last_row_env = os.getenv('EXCEL_LAST_DATA_ROW')
        last_data_row: int
        if last_row_env and last_row_env.strip():
            try:
                last_data_row = int(last_row_env)
                if last_data_row < header_row:
                    raise ValueError(f"EXCEL_LAST_DATA_ROW ({last_data_row}) cannot be before header row ({header_row}).")
            except ValueError as e_last_row: # More specific exception
                raise ValueError(f"Invalid EXCEL_LAST_DATA_ROW value: '{last_row_env}'. Error: {e_last_row}")
        else:
            last_data_row = _find_last_nonempty_row(sheet, name_col_idx + 1, header_row + 1, sheet.max_row)
        
        logger.info(f"[read_subcontractors] Effective last data row: {last_data_row} (Header row: {header_row}, Sheet max row: {sheet.max_row})")
        if last_data_row > sheet.max_row: # Should not happen if _find_last_nonempty_row is correct
            logger.warning(f"Calculated last_data_row ({last_data_row}) exceeds sheet.max_row ({sheet.max_row}). Clamping to sheet.max_row.")
            last_data_row = sheet.max_row
        if last_data_row - header_row > 100000: # Sanity check
            raise ValueError(f"Sanity check: Processing {last_data_row - header_row} rows is too many. Check Excel/config.")

        for r_idx in range(header_row + 1, last_data_row + 1):
            sub_name_val = sheet.cell(row=r_idx, column=name_col_idx + 1).value
            flag_val = sheet.cell(row=r_idx, column=flag_col_idx + 1).value
            
            if not flag_val or str(flag_val).strip().lower() not in ("yes", "y", "true", "1"):
                logger.debug(f"Skipping non-subcontractor row {r_idx} (flag value: '{flag_val}')")
                continue

            if not sub_name_val or not str(sub_name_val).strip():
                logger.warning(f"Skipping row {r_idx} due to empty subcontractor name.")
                continue
            
            sub_name_str = str(sub_name_val).strip()
            sub_id_str: str
            if id_col_idx is not None:
                id_cell_val = sheet.cell(row=r_idx, column=id_col_idx + 1).value
                sub_id_str = str(id_cell_val).strip() if id_cell_val else f"ROW_{r_idx}"
            else:
                sub_id_str = f"ROW_{r_idx}"

            subcontractors.append({'row': r_idx, 'name': sub_name_str, 'id': sub_id_str})
        
        logger.info(f"Read {len(subcontractors)} subcontractors from {excel_path}")
        if not subcontractors:
            logger.warning(f"No subcontractors identified in {excel_path}. Check file, '{subcontractor_name_col_header}' column, and flag column.")
        
        workbook.close() # Close workbook after reading
        return subcontractors, headers

    except FileNotFoundError:
        logger.error(f"Excel file not found at: {excel_path}")
        raise
    except ValueError as ve: # Catch specific ValueErrors from our logic
        logger.error(f"Data or Configuration Error reading Excel {excel_path}: {ve}")
        raise
    except Exception as e: # General catch for other openpyxl or unexpected errors
        logger.error(f"Unexpected error reading Excel file {excel_path}: {e}\n{traceback.format_exc()}")
        raise Exception(f"Error reading Excel file {excel_path}: {e}")

        raise Exception(f"Error reading Excel file {excel_path}: {e}") # Re-raise generic for main handler

def get_output_sheet(workbook: openpyxl.Workbook, output_sheet_name_env: Optional[str] = None) -> Worksheet:
    """
    Returns a worksheet for output. Uses 'SUMMARY' as base.
    If output_sheet_name_env is provided and different from 'SUMMARY', creates a copy.
    Raises error if 'export' sheet exists.
    """
    base_sheet_name = 'SUMMARY'
    if 'export' in workbook.sheetnames: # type: ignore
        raise ValueError("The 'export' worksheet is present. Please remove it. Only 'SUMMARY' is used.")
    if base_sheet_name not in workbook.sheetnames: # type: ignore
        raise ValueError("Could not find 'SUMMARY' worksheet. It is required.")
    
    base_sheet: Worksheet = workbook[base_sheet_name] # type: ignore
    
    # Determine target sheet name from env var or default to SUMMARY
    target_sheet_name = output_sheet_name_env or os.getenv('EXCEL_OUTPUT_SHEET', base_sheet_name)
    
    write_policy_dates_to_target = os.getenv('WRITE_POLICY_DATES', 'false').lower() in ('1', 'true', 'yes', 'on')

    if write_policy_dates_to_target and target_sheet_name != base_sheet_name:
        if target_sheet_name in workbook.sheetnames: # type: ignore
            logger.info(f"Using existing sheet '{target_sheet_name}' for output.")
            return workbook[target_sheet_name] # type: ignore
        else:
            # Create a copy from the base 'SUMMARY' sheet
            new_sheet = workbook.copy_worksheet(base_sheet) # type: ignore
            new_sheet.title = target_sheet_name
            logger.info(f"Created output worksheet '{target_sheet_name}' as a copy of '{base_sheet_name}'.")
            return new_sheet
    else: # Default: use SUMMARY sheet directly, or if not writing policy dates to a separate sheet
        logger.info(f"Using '{base_sheet_name}' worksheet for output.")
        return base_sheet

# def insert_gap_rows(sheet: Worksheet, result: Dict[str, Any], col_map: Dict[str, int]):
#     """
#     For each gap in result['gl_gap_ranges'] and result['wc_gap_ranges'], insert a row below the subcontractor row.
#     No custom formatting is applied to inserted rows (plain rows only).
#     NOTE: This function is currently commented out as its functionality of inserting rows into the SUMMARY
#     sheet conflicts with the stated goal of update_excel to not add gap rows to SUMMARY.
#     If this functionality is desired for a different sheet or context, it should be uncommented and refactored.
#     """
#     row_idx = result.get('row')
#     if not isinstance(row_idx, int): # Ensure row_idx is a valid integer
#         logger.warning(f"insert_gap_rows: Invalid or missing row_idx for result {result.get('name')}. Skipping gap row insertion.")
#         return
    
#     name = result.get('name', '')
#     sub_id = result.get('id', '')
    
#     # Find the name and ID columns from the first row
#     name_col = 1  # Default to column A
#     id_col = 3    # Default to column C
    
#     # Look at headers in row 6 to find the correct columns
#     for col_idx_loop in range(1, 10):  # Check first 10 columns
#         header = sheet.cell(row=6, column=col_idx_loop).value
#         if header and 'name' in str(header).lower():
#             name_col = col_idx_loop
#         if header and 'id' in str(header).lower() and 'subcontractor' in str(header).lower():
#             id_col = col_idx_loop
    
#     gap_types = [('GL', 'gl_gap_ranges'), ('WC', 'wc_gap_ranges')]
#     current_insert_offset = 0 # Keep track of how many rows we've inserted for this sub
#     for gap_type, key in gap_types:
#         for gap in result.get(key, []):
#             actual_insert_row = row_idx + 1 + current_insert_offset
#             sheet.insert_rows(actual_insert_row)
            
#             sheet.cell(row=actual_insert_row, column=name_col).value = name
#             if id_col:
#                 sheet.cell(row=actual_insert_row, column=id_col).value = sub_id
            
#             if gap_type == 'GL':
#                 sheet.cell(row=actual_insert_row, column=col_map[GL_FROM_KEY]).value = gap[0]
#                 sheet.cell(row=actual_insert_row, column=col_map[GL_TO_KEY]).value = gap[1]
#                 if NOTES_HEADER in col_map:
#                     sheet.cell(row=actual_insert_row, column=col_map[NOTES_HEADER]).value = "GL Coverage Gap"
#             elif gap_type == 'WC':
#                 sheet.cell(row=actual_insert_row, column=col_map[WC_FROM_KEY]).value = gap[0]
#                 sheet.cell(row=actual_insert_row, column=col_map[WC_TO_KEY]).value = gap[1]
#                 if NOTES_HEADER in col_map:
#                     sheet.cell(row=actual_insert_row, column=col_map[NOTES_HEADER]).value = "WC Coverage Gap"
            
#             current_insert_offset += 1

def update_excel(excel_path_str: str, results: List[Dict[str, Any]], headers: List[Any]):
    """Updates the Excel file with audit results. Writes policy dates to the target summary sheet."""
    excel_path = Path(excel_path_str)
    try:
        workbook = openpyxl.load_workbook(excel_path)
        
        # Determine target sheet for writing policy dates
        # The get_output_sheet function handles creation if EXCEL_OUTPUT_SHEET is different from SUMMARY
        sheet = get_output_sheet(workbook) # output_sheet_name_env is implicitly from os.getenv in get_output_sheet
        
        # Fixed column mapping for GL and WC dates on the SUMMARY sheet (or its copy)
        # These are 1-based column indices.
        col_map = {
            GL_FROM_KEY: 9,    # Column I
            GL_TO_KEY: 10,   # Column J
            WC_FROM_KEY: 11,  # Column K
            WC_TO_KEY: 12,   # Column L
        }
        logger.info(f"Using fixed column mapping for date updates on sheet '{sheet.title}': GL From=I(9), GL To=J(10), WC From=K(11), WC To=L(12)")
        
        update_count = 0
        # Sort by row index to process consistently, though not strictly necessary for this update logic.
        for result in sorted(results, key=lambda x: x.get('row', float('inf'))):
            row_idx = result.get('row')
            if not row_idx or not isinstance(row_idx, int):
                logger.warning(f"Skipping result with missing or invalid 'row' identifier: {result.get('name', 'N/A')}")
                continue

            def format_date_cell_for_excel(dt_val: Any) -> Optional[dt.datetime]:
                """Converts date to datetime for Excel compatibility, returns None otherwise."""
                if isinstance(dt_val, dt.datetime):
                    return dt_val
                if isinstance(dt_val, dt.date):
                    return dt.datetime.combine(dt_val, dt.time.min) # Convert date to datetime
                return None

            try:
                date_columns_keys = [GL_FROM_KEY, GL_TO_KEY, WC_FROM_KEY, WC_TO_KEY]
                date_values_map = {
                    GL_FROM_KEY: result.get('gl_from'),
                    GL_TO_KEY: result.get('gl_to'),
                    WC_FROM_KEY: result.get('wc_from'),
                    WC_TO_KEY: result.get('wc_to')
                }

                for key_col in date_columns_keys:
                    # Get the cell object from the sheet.
                    target_cell = sheet.cell(row=row_idx, column=col_map[key_col])
                    original_cell_coordinate = target_cell.coordinate
                    
                    was_merged_and_unmerged = False
                    # Iterate over a copy of merged_cells.ranges because unmerging modifies the collection.
                    for merged_cell_range in list(sheet.merged_cells.ranges):
                        if original_cell_coordinate in merged_cell_range:
                            logger.debug(f"Cell {original_cell_coordinate} is part of merged range {merged_cell_range}. Unmerging.")
                            sheet.unmerge_cells(str(merged_cell_range))
                            was_merged_and_unmerged = True
                            # After unmerging, the original cell object `target_cell` should now refer to the
                            # top-left cell of the unmerged area. Re-fetching is good practice.
                            target_cell = sheet.cell(row=row_idx, column=col_map[key_col])
                            break
                    
                    value_to_assign = format_date_cell_for_excel(date_values_map[key_col])
                    target_cell.value = value_to_assign
                    
                    if value_to_assign is not None:
                        target_cell.number_format = 'YYYY-MM-DD'
                    elif was_merged_and_unmerged: # If unmerged and value is now None
                        target_cell.number_format = 'General' # Reset format if it was a date before
                    # If not merged and value is None, cell is simply cleared, existing format might persist or be irrelevant.



                # Note: The original logic for GL_GAP_COL, WC_GAP_COL, NOTES_COL was removed
                # as it seemed to conflict with the "Only map the fixed columns for SUMMARY" comment.
                # If these need to be written to the *same* sheet, their column indices need to be defined.
                # The current structure implies they are for a separate GAPS_REPORT/ERRORS_REPORT.

                update_count += 1
            except Exception as cell_e:
                logger.error(f"Error updating cell for row {row_idx}, sub '{result.get('name', 'N/A')}': {cell_e}", exc_info=True)
        
        try:
            workbook.save(excel_path)
            logger.info(f"Successfully updated {update_count} rows in sheet '{sheet.title}' of {excel_path}")
        except PermissionError:
            logger.error(f"Permission denied saving Excel file: {excel_path}. Ensure the file is not open.")
            raise # Re-raise to be caught by main
        except Exception as save_e:
            logger.error(f"Failed to save Excel file {excel_path}: {save_e}")
            raise # Re-raise
        finally:
            workbook.close() # Ensure workbook is closed

    except Exception as e:
        logger.error(f"Error processing or updating Excel file {excel_path}: {e}")
        raise

        raise # Re-raise to be caught by main

def _apply_basic_formatting(worksheet: Optional[Worksheet], header_row_idx: int = 1):
    """Applies basic formatting: bold header, auto-fit columns, freeze header row."""
    if not worksheet:
        logger.debug(f"Skipping formatting for None worksheet object.")
        return

    sheet_title = worksheet.title if hasattr(worksheet, 'title') else "UnknownSheet"
    logger.debug(f"Applying basic formatting to worksheet: '{sheet_title}' with header row {header_row_idx}.")

    bold_font = Font(bold=True)
    if worksheet.max_row >= header_row_idx and worksheet.max_column >= 1:
        try:
            if header_row_idx <= worksheet.max_row: # Ensure row exists
                for cell in worksheet[header_row_idx]:
                    if cell: cell.font = bold_font
                logger.debug(f"Applied bold font to header row {header_row_idx} for sheet '{sheet_title}'.")
            else:
                logger.warning(f"Header row {header_row_idx} out of bounds (max_row: {worksheet.max_row}) for sheet '{sheet_title}'. Skipping bolding.")
        except Exception as e_font:
            logger.warning(f"Could not apply bold font to header row {header_row_idx} for sheet '{sheet_title}': {e_font}", exc_info=True)

    max_lengths: Dict[int, int] = {}
    try:
        for col_idx_iter, column_iter_obj in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), 1):
            current_max_len = 0
            for cell_obj in column_iter_obj:
                if cell_obj.value is not None:
                    try:
                        cell_val_str = str(cell_obj.value)
                        current_max_len = max(current_max_len, len(cell_val_str))
                    except Exception as cell_str_e:
                        logger.debug(f"Could not convert cell value to string for length calc in sheet '{sheet_title}', col {col_idx_iter}: {cell_str_e}")
            if current_max_len > 0:
                max_lengths[col_idx_iter] = current_max_len
    except Exception as e_width:
        logger.warning(f"Error during column width calculation for sheet '{sheet_title}': {e_width}", exc_info=True)

    if max_lengths:
        for col_idx_map, length_val in max_lengths.items():
            column_letter = get_column_letter(col_idx_map)
            adjusted_width = min(length_val + 4, 75) # Increased padding slightly more
            worksheet.column_dimensions[column_letter].width = adjusted_width
        logger.debug(f"Auto-fitted column widths for sheet '{sheet_title}'.")
    else:
        logger.debug(f"No content found to auto-fit column widths for sheet '{sheet_title}'.")

    if worksheet.max_row > header_row_idx and header_row_idx >= 1:
        try:
            pane_to_freeze_at = worksheet.cell(row=header_row_idx + 1, column=1)
            worksheet.freeze_panes = pane_to_freeze_at.coordinate
            logger.debug(f"Froze header row at {pane_to_freeze_at.coordinate} for sheet '{sheet_title}'.")
        except Exception as e_freeze:
            logger.warning(f"Could not freeze panes for sheet '{sheet_title}': {e_freeze}", exc_info=True)
    elif header_row_idx == 0:
        logger.debug(f"Skipping freeze panes for sheet '{sheet_title}' due to invalid header_row_idx: {header_row_idx}.")
    else: # max_row <= header_row_idx or sheet is empty/only header
        logger.debug(f"Skipping freeze panes for sheet '{sheet_title}' as it only contains the header row or is empty.")

def write_gaps_report(output_dir_str: str, gaps: List[Dict[str, Any]], excel_path_str: Optional[str] = None):
    """Writes the gap report to a CSV file, and GAPS_REPORT/ERRORS_REPORT worksheets to Excel if excel_path_str is provided."""
    output_dir = Path(output_dir_str)
    report_path_csv = output_dir / 'gaps_report.csv'
    
    if not gaps:
        logger.info("No gaps or issues found to report. Skipping CSV and worksheet generation.")
        return

    try:
        # Define a more comprehensive set of standard fieldnames for CSV
        fieldnames_std_csv = [
            'subcontractor_name', 'subcontractor_id', 'issue_type', 'details',
            'policy_type', 'effective_date', 'expiration_date',
            'gap_start', 'gap_end', 'file_path', 'page_number'
        ]
        all_actual_keys = set().union(*(d.keys() for d in gaps))
        # Ensure standard fields come first, then any others, without duplicates
        fieldnames_for_csv = list(dict.fromkeys(fieldnames_std_csv + sorted(list(all_actual_keys - set(fieldnames_std_csv)))))
        
        with open(report_path_csv, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames_for_csv, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(gaps)
        logger.info(f"Gaps report CSV saved to {report_path_csv}")
    except PermissionError:
        logger.error(f"Permission denied writing gaps report CSV: {report_path_csv}. Check file permissions or if it's open.")
        raise
    except Exception as e_csv:
        logger.error(f"Error writing gaps report CSV to {report_path_csv}: {e_csv}")
        raise Exception(f"Error writing gaps report CSV to {report_path_csv}: {e_csv}")

    if excel_path_str:
        excel_path_obj = Path(excel_path_str)
        workbook: Optional[openpyxl.Workbook] = None # Initialize to None
        try:
            workbook = openpyxl.load_workbook(excel_path_obj)
            
            # GAPS_REPORT sheet
            gap_rows_data = [g for g in gaps if g.get('issue_type') == STATUS_GAP] # Use constant
            gap_fields_excel = ['subcontractor_name', 'policy_type', 'effective_date', 'expiration_date', 'details', 'file_path'] # More relevant fields for gaps
            
            if 'GAPS_REPORT' in workbook.sheetnames: # type: ignore
                del workbook['GAPS_REPORT'] # type: ignore
            ws_gaps_report: Worksheet = workbook.create_sheet('GAPS_REPORT') # type: ignore
            for c_idx, field_name in enumerate(gap_fields_excel, 1):
                ws_gaps_report.cell(row=1, column=c_idx, value=field_name)
            for r_idx, gap_item in enumerate(gap_rows_data, 2):
                for c_idx, field_name in enumerate(gap_fields_excel, 1):
                    ws_gaps_report.cell(row=r_idx, column=c_idx, value=gap_item.get(field_name, ''))
            _apply_basic_formatting(ws_gaps_report, header_row_idx=1)
            logger.info(f"GAPS_REPORT worksheet created with {len(gap_rows_data)} gap rows.")

            # ERRORS_REPORT sheet (all non-Gap issues)
            error_rows_data = [g for g in gaps if g.get('issue_type') != STATUS_GAP]
            if 'ERRORS_REPORT' in workbook.sheetnames: # type: ignore
                del workbook['ERRORS_REPORT'] # type: ignore
            
            if error_rows_data:
                ws_errors_report: Worksheet = workbook.create_sheet('ERRORS_REPORT') # type: ignore
                # Use a comprehensive set of headers for errors, similar to CSV, excluding pure gap fields
                error_report_headers_excel = [
                    'subcontractor_name', 'subcontractor_id', 'issue_type', 'details',
                    'policy_type', 'effective_date', 'expiration_date', 'file_path', 'page_number'
                ]
                # Filter to keys actually present in error_rows_data to avoid empty columns if some keys are always missing
                actual_error_keys = set().union(*(d.keys() for d in error_rows_data))
                final_error_headers = [h for h in error_report_headers_excel if h in actual_error_keys]
                if not final_error_headers and error_rows_data : # Fallback if predefined headers don't match data
                    final_error_headers = sorted(list(actual_error_keys))


                for c_idx, field_name in enumerate(final_error_headers, 1):
                    ws_errors_report.cell(row=1, column=c_idx, value=field_name)
                for r_idx, err_item in enumerate(error_rows_data, 2):
                    for c_idx, field_name in enumerate(final_error_headers, 1):
                        ws_errors_report.cell(row=r_idx, column=c_idx, value=err_item.get(field_name, ''))
                _apply_basic_formatting(ws_errors_report, header_row_idx=1)
                logger.info(f"ERRORS_REPORT worksheet created with {len(error_rows_data)} error/issue rows.")
            else:
                logger.info("No non-gap errors to report in ERRORS_REPORT worksheet.")

            workbook.save(excel_path_obj)
            logger.info(f"Gaps and errors worksheets written to {excel_path_obj}")
        except Exception as e_excel_report:
            logger.error(f"Failed to write GAPS_REPORT or ERRORS_REPORT worksheet to {excel_path_obj}: {e_excel_report}")
            raise Exception(f"Failed to write GAPS_REPORT or ERRORS_REPORT worksheet: {e}")


            raise # Re-raise to be caught by main
        finally:
            if 'workbook' in locals() and workbook: workbook.close()


def run_excel_sanity_checks(excel_path_str: Optional[str] = None):
    """Run basic sanity checks on Excel file and config. Abort if failed."""
    import sys # Keep local for sys.exit
    import traceback # Keep local for traceback formatting

    final_excel_path_str = excel_path_str or os.getenv('EXCEL_FILE_PATH')
    
    if not final_excel_path_str:
        logger.critical("Sanity check failed: EXCEL_FILE_PATH not provided or found in .env.")
        print("Sanity check failed: EXCEL_FILE_PATH not configured.")
        sys.exit(1)
        
    final_excel_path = Path(final_excel_path_str)
    if not final_excel_path.exists():
        logger.critical(f"Sanity check failed: Excel file not found at {final_excel_path}")
        print(f"Sanity check failed: Excel file not found at {final_excel_path}")
        sys.exit(1)
    try:
        logger.info(f"Running sanity checks on Excel file: {final_excel_path}")
        # read_subcontractors now expects a string path
        read_subcontractors(str(final_excel_path))
        logger.info("Excel sanity checks passed.")
        print("Excel sanity checks passed.")
    except Exception as e_sanity:
        logger.critical(f"Sanity check failed for {final_excel_path}: {e_sanity}\n{traceback.format_exc()}")
        print(f"Sanity check failed for {final_excel_path}: {e_sanity}\n{traceback.format_exc()}")
        sys.exit(1)

def create_error_report_workbook_for_validation(output_path: Path, errors_data: List[Dict[str, Any]], headers: List[str]):
    """
    Creates a new Excel workbook with an "ERROR_REPORT" sheet populated with
    the given errors_data and headers. This is used for the --validate-batch feature.

    Args:
        output_path: The Path where the Excel file will be saved.
        errors_data: A list of dictionaries, where each dictionary represents an error row.
                     The keys in the dictionary should correspond to the provided headers.
        headers: A list of strings representing the column headers for the "ERROR_REPORT" sheet.
    """
    logging.info(f"Creating error report workbook for validation at: {output_path}")
    workbook = openpyxl.Workbook()

    # Remove default sheet if it exists (usually named "Sheet")
    if "Sheet" in workbook.sheetnames:
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)
        logging.debug("Removed default 'Sheet' from new workbook.")

    error_sheet = workbook.create_sheet("ERROR_REPORT")
    logging.debug("Created 'ERROR_REPORT' sheet.")

    # Write headers
    for col_idx, header_title in enumerate(headers, 1):
        error_sheet.cell(row=1, column=col_idx, value=header_title)
    logging.debug(f"Wrote headers: {headers}")

    # Write error data
    for row_idx, error_item in enumerate(errors_data, 2): # Start data from row 2
        for col_idx, header_key in enumerate(headers, 1):
            cell_value = error_item.get(header_key, "") # Get value by header key
            error_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    _apply_basic_formatting(error_sheet, header_row_idx=1) # Apply formatting
    logging.debug(f"Wrote {len(errors_data)} error rows to 'ERROR_REPORT'.")

    try:
        workbook.save(str(output_path)) # save expects string path
        logger.info(f"Successfully saved validation error report to: {output_path}")
    except PermissionError:
        logger.error(f"Permission denied saving validation error report: {output_path}. Ensure the file is not open and permissions are correct.")
        raise
    except Exception as e_save_val_report:
        logger.error(f"Error saving validation error report to {output_path}: {e_save_val_report}")
        raise
    finally:
        if 'workbook' in locals() and workbook: workbook.close()
        logging.error(f"Failed to save validation error report {output_path}: {e}", exc_info=True)
        # Re-raise
        raise


if __name__ == '__main__':
    # Basic test for running module directly (limited without actual files)
    print("Excel handler module - run main.py for full execution.")
    logging.basicConfig(level=logging.INFO) # Ensure logging works when run directly
    print(f"Expected input column: {SUBCONTRACTOR_ID_COL}")
    print(f"Optional input column: {SUBCONTRACTOR_ID_COL}")
    # print(f"Output columns: {', '.join(OUTPUT_COLUMNS)}") # OUTPUT_COLUMNS was not defined, commented out
    
    # Test for create_error_report_workbook_for_validation
    print("\nTesting create_error_report_workbook_for_validation...")
    test_headers = ["Col A", "Col B", "Col C"]
    test_data = [
        {"Col A": "Data1A", "Col B": "Data1B", "Col C": "Data1C"},
        {"Col A": "Data2A", "Col B": "Data2B", "Col C": "Data2C", "ExtraCol": "Should be ignored"},
        {"Col B": "Data3BOnly"} # Missing Col A and Col C
    ]
    test_output_path = Path("temp_validation_report_test.xlsx")
    try:
        create_error_report_workbook_for_validation(test_output_path, test_data, test_headers)
        # Verify content (basic check)
        wb_check = openpyxl.load_workbook(test_output_path)
        assert "ERROR_REPORT" in wb_check.sheetnames
        sheet_check = wb_check["ERROR_REPORT"]
        assert sheet_check.cell(row=1, column=1).value == "Col A"
        assert sheet_check.cell(row=2, column=1).value == "Data1A"
        assert sheet_check.cell(row=4, column=1).value is None # For {"Col B": "Data3BOnly"}
        assert sheet_check.cell(row=4, column=2).value == "Data3BOnly"
        print(f"Test report created at {test_output_path}. Please inspect manually.")
        # test_output_path.unlink() # Clean up
        print(f"Test successful. Cleaned up {test_output_path} (manual deletion might be needed if open).")

    except Exception as e:
        print(f"Test failed: {e}")
    finally:
        if test_output_path.exists():
             try:
                test_output_path.unlink() # Ensure cleanup
             except PermissionError:
                print(f"Could not delete {test_output_path} as it might be open.")
